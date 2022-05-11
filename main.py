import os.path

import fitz
import cv2
from os import *
import time
import openpyxl


urls = []

#Abrimos excel
bk = openpyxl.load_workbook("urls.xlsx")
ex = bk.active


def getImages(doc):
    global urls
    for i in range(len(doc)):
        if i == 1:
            for img in doc.get_page_images(i):
                xref = img[0]
                if xref == 8 or xref == 12:
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n < 5:       # this is GRAY or RGB
                        pix.save("image.png")
                        image = cv2.imread("image.png")
                        decoder = cv2.QRCodeDetector()
                        decodedText, points, _ = decoder.detectAndDecode(image)
                        if points is not None:
                            urls.append(decodedText)
                            remove("image.png")
                            return decodedText
                        else:
                            pass


                    else:               # CMYK: convert to RGB first
                        pix1 = fitz.Pixmap(fitz.csRGB, pix)
                        pix1.save("p%s-%s.png" % (i, xref))
                        pix1 = None
                    pix = None


files = listdir("PDFS")
i = 1
for file in files:
    found = False
    document = os.path.join('PDFS', file)
    doc = fitz.Document(document)
    url = getImages(doc)
    urlsCell = ex.cell(row=i, column=1)
    urlsCell.value = url
    nameCell = ex.cell(row=i, column=2)
    nameCell.value = file
    i += 1

for i in range(len(urls)):
    print(i, ") ", urls[i])

bk.save('urls.xlsx')